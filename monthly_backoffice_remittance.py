#!/usr/bin/env python3
"""Monthly Back-Office Fee Remittance — Hetzner (tourscale-reports) edition.

For a target month it:
  1. Pulls the back-office Peek fee (totalPeekFees on Peek Pro Mobile/Web display
     sources) per location via the peek-app GraphQL proxy, purchase-date basis.
  2. Updates the per-entity YTD spreadsheet (Jan..target month).
  3. Issues the per-entity monthly Remittance-Statement PDF (skips $0 months).
  4. Builds the Ramp Bill Pay manifest for the month.
  5. NOTIFY: emails the operator a review summary + attachments (no Ramp send).
     SEND:   emails each PDF to the Ramp bills inbox (one email = one bill).

Usage (via scripts/run.sh so .env is sourced):
    scripts/run.sh monthly_backoffice_remittance.py [YYYY-MM] [--notify|--send]
    (no month -> previous completed calendar month)

Env (.env): PEEK_APP_INTERNAL_TOKEN, RAMP_BILLS_EMAIL, REMITTANCE_NOTIFY_EMAIL,
REPORTS_SMTP_* / REPORTS_FROM_EMAIL (shared), CHROMIUM_BIN (default /usr/bin/chromium).
The 6% back-office fee = Peek's totalPeekFees on Peek Pro Mobile/Web — see the
'backoffice-6pct-fee-remittance' note in the royalties worksheet.
"""
from __future__ import annotations
import os, sys, json, csv, base64, calendar, subprocess, tempfile, shutil, urllib.request
from datetime import date

from lib.email import send as send_email

ROOT = os.path.dirname(os.path.abspath(__file__))
APP = "https://peek-app.tourscale.com"
LOGO = os.path.join(ROOT, "assets", "tourscale-logo-dark.svg")
CHROME = os.environ.get("CHROMIUM_BIN", "/usr/bin/chromium")
OUTBASE = os.path.join(ROOT, "output", "backoffice_fee")
BACKOFFICE = {"Peek Pro Mobile", "Peek Pro Web", "peek_pro_web", "peek_pro_mobile"}

RAMP_BILLS_EMAIL = os.environ.get("RAMP_BILLS_EMAIL", "tourscalefranchising@ap.ramp.com")
NOTIFY_EMAIL = os.environ.get("REMITTANCE_NOTIFY_EMAIL", "kai@tourscale.com")
SEND_CMD_HINT = "/opt/tourscale/reports/scripts/run.sh monthly_backoffice_remittance.py"

ISSUER = {
    "name": "TourScale Enterprises, LLC",
    "addr": ["1300 South Blvd, Suite 30102", "Charlotte, NC 28203"],
    "phone": "888-753-7507", "email": "finance@tourscale.com", "web": "www.tourscale.com",
}
ENTITIES = [
    {"code": "JAB", "name": "JAB Boating, LLC",
     "addr": ["260 Braid Dr", "Mt. Juliet, TN 37122"], "email": "nashville@cruisintikis.com",
     "locations": [("Cruisin' Tikis Nashville", "cruisin-tikis-nashville")]},
    {"code": "TTR", "name": "Tiki Times Rentals, LLC",
     "addr": ["127 SE 31st St", "Cape Coral, FL 33904"], "email": "capecoral@cruisintikis.com",
     "locations": [("Cruisin' Tikis Cape Coral", "cruisin-tikis-cape-coral"),
                   ("Cruisin' Tikis Fort Myers Beach", "cruisin-tikis-fort-myers-beach"),
                   ("Cruisin' Tikis St James City", "cruisin-tikis-st-james-city")]},
    {"code": "CCE", "name": "Cape Coral Entertainment LLC",
     "addr": ["127 SE 31st St", "Cape Coral, FL 33904"], "email": "nick@trolleypub.com",
     "locations": [("Trolley Pub Cape Coral", "trolley-pub-cape-coral")]},
]

# ---------------- Peek pull ----------------
_Q = ("query G($f:[AggregateAmountsFilter!]!,$c:Currency!,$g:[AggregateAmountsGrouping!]!){"
      "queryAggregateAmounts(currency:$c,filters:$f,groupings:$g){__typename"
      " ... on AggregateAmountsSuccess{aggregateAmounts{aggregateAmounts{"
      "displaySource bookingCount priceRetail totalPeekFees}}}"
      " ... on GenericError{__typename} ... on NoResultsError{__typename}}}")

def token():
    t = os.environ.get("PEEK_APP_INTERNAL_TOKEN", "").strip()
    if not t:
        sys.exit("PEEK_APP_INTERNAL_TOKEN not set (add it to .env)")
    return t

def pull(slug, ym, tok):
    y, m = int(ym[:4]), int(ym[5:7]); end = f"{ym}-{calendar.monthrange(y, m)[1]:02d}"
    v = {"c": "USD", "g": ["DISPLAY_SOURCE"], "f": [
        {"key": "BOOKED_ON_DATE", "operation": "GTE", "value": {"date": f"{ym}-01"}},
        {"key": "BOOKED_ON_DATE", "operation": "LTE", "value": {"date": end}}]}
    req = urllib.request.Request(f"{APP}/api/installs/{slug}/peek-graphql",
        data=json.dumps({"query": _Q, "variables": v}).encode(), method="POST",
        headers={"Authorization": f"Bearer {tok}", "Content-Type": "application/json"})
    n = (json.loads(urllib.request.urlopen(req, timeout=120).read()).get("data") or {}).get("queryAggregateAmounts") or {}
    rows = (n.get("aggregateAmounts") or {}).get("aggregateAmounts") or []
    bo = [r for r in rows if r.get("displaySource") in BACKOFFICE]
    return {"bookings": sum(int(r.get("bookingCount") or 0) for r in bo),
            "retail": round(sum(float(r.get("priceRetail") or 0) for r in bo), 2),
            "fee": round(sum(float(r.get("totalPeekFees") or 0) for r in bo), 2)}

def gather(year, thru_month, tok):
    months = [f"{year}-{m:02d}" for m in range(1, thru_month + 1)]
    data = {}
    for ent in ENTITIES:
        for _disp, slug in ent["locations"]:
            data[slug] = {ym: pull(slug, ym, tok) for ym in months}
    return data, months

def month_label(ym):
    y, m = int(ym[:4]), int(ym[5:7]); return f"{calendar.month_name[m]} {y}"

# ---------------- PDF ----------------
def _logo_uri():
    return "data:image/svg+xml;base64," + base64.b64encode(open(LOGO, "rb").read()).decode()

_CSS = """
@page { size: letter; margin: 0.7in 0.7in; } * { box-sizing:border-box; }
body { font-family:'Helvetica Neue',Arial,sans-serif; color:#1a1a1a; margin:0; font-size:12px; }
.top { display:flex; justify-content:space-between; align-items:flex-start; border-bottom:3px solid #1F3864; padding-bottom:16px; }
.logo img { height:42px; } .issuer { text-align:right; font-size:10.5px; color:#444; line-height:1.5; }
.issuer .nm { font-weight:700; color:#1F3864; font-size:12.5px; }
h1 { color:#1F3864; font-size:22px; letter-spacing:.5px; margin:26px 0 2px; font-weight:800; }
.sub { color:#7a7a7a; font-size:11px; margin-bottom:22px; }
.meta { display:flex; justify-content:space-between; margin-bottom:26px; gap:24px; } .box { flex:1; }
.box .lbl { text-transform:uppercase; letter-spacing:1px; font-size:9px; color:#9a9a9a; margin-bottom:6px; font-weight:700; }
.box .val { font-size:11.5px; line-height:1.6; } .box .val .strong { font-weight:700; color:#1F3864; font-size:12.5px; }
table { width:100%; border-collapse:collapse; margin-top:6px; }
th { background:#1F3864; color:#fff; font-size:10px; letter-spacing:.6px; text-transform:uppercase; padding:10px 12px; text-align:left; }
th.r,td.r { text-align:right; } th.c,td.c { text-align:center; }
td { padding:11px 12px; border-bottom:1px solid #e6e6e6; font-size:11.5px; } tr:nth-child(even) td { background:#f7f8fb; }
.tot td { border-top:2px solid #1F3864; border-bottom:none; font-weight:800; font-size:13px; color:#1F3864; background:#eaf0fa !important; padding:13px 12px; }
.remit { margin-top:30px; background:#E2EFDA; border-left:4px solid #4a7d3a; padding:16px 18px; border-radius:3px; display:flex; justify-content:space-between; align-items:center; }
.remit .t { font-size:11px; text-transform:uppercase; letter-spacing:1px; color:#3d5c2f; font-weight:700; } .remit .a { font-size:26px; font-weight:800; color:#2f4a24; }
.note { margin-top:26px; font-size:9.5px; color:#8a8a8a; line-height:1.6; border-top:1px solid #eee; padding-top:12px; }
.foot { margin-top:14px; font-size:9px; color:#b0b0b0; text-align:center; }
"""

def build_html(ent, ym, rows, total, issue_date):
    multi = len(rows) > 1
    loc_rows = "".join(
        f"<tr><td>Peek back-office booking fee &mdash; {d}</td><td class='c'>{b}</td>"
        f"<td class='r'>${rt:,.2f}</td><td class='r'>${fe:,.2f}</td></tr>" for d, b, rt, fe in rows)
    total_row = (f"<tr class='tot'><td>Total remittance &mdash; {month_label(ym)}</td>"
                 f"<td class='c'></td><td class='r'></td><td class='r'>${total:,.2f}</td></tr>") if multi else ""
    stmt_no = f"BOFR-{ent['code']}-{ym.replace('-','')}"
    return f"""<!doctype html><html><head><meta charset='utf-8'><style>{_CSS}</style></head><body>
<div class='top'><div class='logo'><img src='{_logo_uri()}'></div>
<div class='issuer'><div class='nm'>{ISSUER['name']}</div>{'<br>'.join(ISSUER['addr'])}<br>{ISSUER['phone']} &middot; {ISSUER['email']}<br>{ISSUER['web']}</div></div>
<h1>Back-Office Fee Remittance</h1>
<div class='sub'>Remittance of the 6% back-office booking fee charged by Peek on phone / staff-entered bookings</div>
<div class='meta'>
  <div class='box'><div class='lbl'>Remit To</div><div class='val'><span class='strong'>{ent['name']}</span><br>{'<br>'.join(ent['addr'])}<br>{ent['email']}</div></div>
  <div class='box'><div class='lbl'>Statement Details</div><div class='val'>
     <b>Statement No:</b> {stmt_no}<br><b>Service Period:</b> {month_label(ym)}<br>
     <b>Statement Date:</b> {issue_date}<br><b>Basis:</b> Booking purchase date &middot; USD</div></div></div>
<table><thead><tr><th>Description</th><th class='c'>Back-office bookings</th><th class='r'>Back-office sales (list price)</th><th class='r'>Fee remitted</th></tr></thead>
<tbody>{loc_rows}{total_row}</tbody></table>
<div class='remit'><div class='t'>Total Remitted to Franchisee</div><div class='a'>${total:,.2f}</div></div>
<div class='note'>This statement reflects the 6% back-office booking fee that Peek charges on bookings placed through the back office (over the phone or entered by staff) rather than online, which TourScale Enterprises, LLC remits back to the Franchisee per agreement. The fee amount is the actual back-office fee recorded by Peek for the service period, bucketed by booking purchase date. Back-office bookings are those originating from Peek Pro (Mobile / Web); online (widget) and third-party / OTA bookings are excluded.</div>
<div class='foot'>Generated {issue_date} &middot; {ISSUER['name']} &middot; {stmt_no}</div></body></html>"""

def render_pdf(html, out_pdf):
    tmp = tempfile.mkdtemp(prefix="ts_bofr_")
    try:
        cp = os.path.join(tmp, "s.html"); open(cp, "w").write(html)
        subprocess.run([CHROME, "--headless=new", "--disable-gpu", "--no-sandbox",
            "--virtual-time-budget=6000", "--print-to-pdf-no-header",
            f"--print-to-pdf={out_pdf}", f"file://{cp}"], capture_output=True, timeout=90)
        return os.path.exists(out_pdf) and os.path.getsize(out_pdf) > 1000
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

# ---------------- spreadsheet (per entity, YTD) ----------------
def update_workbook(ent, data, months):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    navy, grey, green = "1F3864", "F2F2F2", "E2EFDA"
    thin = Side(style="thin", color="BFBFBF"); B = Border(thin, thin, thin, thin)
    def hdr(c): c.font = Font(bold=True, color="FFFFFF", size=10); c.fill = PatternFill("solid", fgColor=navy); c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = B
    def money(c): c.number_format = '#,##0.00'; c.alignment = Alignment(horizontal="right")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Back-Office Fee"; ws.sheet_view.showGridLines = False
    locs = ent["locations"]
    ws["A1"] = f"{ent['name']} — Back-Office Fee Remittance (YTD)"; ws["A1"].font = Font(bold=True, size=14, color=navy)
    ws["A2"] = "6% Peek back-office booking fee (phone/staff) · recorded totalPeekFees · purchase-date basis · USD"
    ws["A2"].font = Font(italic=True, size=9, color="808080")
    r0 = 4
    hdr(ws.cell(row=r0, column=1, value="Month"))
    for j, (disp, _) in enumerate(locs, start=2): hdr(ws.cell(row=r0, column=j, value=disp))
    if len(locs) > 1: hdr(ws.cell(row=r0, column=2 + len(locs), value="Monthly Total ($)"))
    col_tot = [0.0] * len(locs); grand = 0.0
    for i, ym in enumerate(months):
        r = r0 + 1 + i
        mc = ws.cell(row=r, column=1, value=month_label(ym)); mc.border = B
        row_tot = 0.0
        for j, (disp, slug) in enumerate(locs, start=2):
            fee = data[slug][ym]["fee"]; c = ws.cell(row=r, column=j, value=fee); money(c); c.border = B
            col_tot[j - 2] += fee; row_tot += fee
        if len(locs) > 1:
            tc = ws.cell(row=r, column=2 + len(locs), value=round(row_tot, 2)); money(tc); tc.border = B; tc.font = Font(bold=True)
        grand += row_tot
        if i % 2:
            for j in range(1, 2 + len(locs) + (1 if len(locs) > 1 else 0)): ws.cell(row=r, column=j).fill = PatternFill("solid", fgColor=grey)
    r = r0 + 1 + len(months)
    yc = ws.cell(row=r, column=1, value="YTD Total"); yc.font = Font(bold=True); yc.border = B
    for j in range(len(locs)):
        c = ws.cell(row=r, column=2 + j, value=round(col_tot[j], 2)); money(c); c.font = Font(bold=True); c.border = B
    if len(locs) > 1:
        gc = ws.cell(row=r, column=2 + len(locs), value=round(grand, 2)); money(gc); gc.font = Font(bold=True, size=12); gc.border = B
    for j in range(1, 2 + len(locs) + (1 if len(locs) > 1 else 0)): ws.cell(row=r, column=j).fill = PatternFill("solid", fgColor=green)
    ws.column_dimensions["A"].width = 16
    for j in range(2, 2 + len(locs) + (1 if len(locs) > 1 else 0)): ws.column_dimensions[get_column_letter(j)].width = 26
    safe = ent["name"].replace(",", "").replace(" ", "_")
    d = os.path.join(OUTBASE, ent["name"]); os.makedirs(d, exist_ok=True)
    path = os.path.join(d, f"{safe}_BackOffice_Fee_YTD_{months[-1]}.xlsx")
    wb.save(path)
    return path, round(grand if len(locs) > 1 else col_tot[0], 2)

# ---------------- manifest ----------------
def write_manifest(bills, ym):
    cols = ["Vendor (legal entity)", "Locations covered", "Bill / Invoice No", "Invoice Date",
            "Due Date", "Terms", "Currency", "Amount (USD)", "Memo", "GL Account / Category", "Attachment (PDF)"]
    os.makedirs(OUTBASE, exist_ok=True)
    path = os.path.join(OUTBASE, f"Ramp_BillPay_Manifest_{ym}.csv")
    with open(path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols); w.writeheader()
        for b in bills: w.writerow({c: b.get(c, "") for c in cols})
    return path

# ---------------- email ----------------
def notify_summary(bills, manifest_path, ym, already_sent):
    total = sum(b["Amount (USD)"] for b in bills)
    lines = "".join(f"<tr><td>{b['Vendor (legal entity)']}</td><td>{b['Bill / Invoice No']}</td>"
                    f"<td style='text-align:right'>${b['Amount (USD)']:,.2f}</td></tr>" for b in bills)
    if already_sent:
        status = (f"<p>Each bill was emailed to the Ramp bills inbox (<b>{RAMP_BILLS_EMAIL}</b>) with you CC'd — "
                  f"review &amp; approve the drafts in Ramp Bill Pay. The Ramp manifest is attached.</p>")
        tag = "[Filed]"
    else:
        send_cmd = f"ssh hetzner '{SEND_CMD_HINT} {ym} --send'"
        status = (f"<p><b>To send these bills to Ramp</b> (emails each PDF to {RAMP_BILLS_EMAIL}, CC you):</p>"
                  f"<pre style='background:#f2f2f2;padding:10px;border-radius:4px'>{send_cmd}</pre>"
                  f"<p style='color:#888;font-size:11px'>Attached: Ramp manifest + the remittance PDFs. Nothing sent to Ramp yet.</p>")
        tag = "[Review]"
    html = (f"<div style='font-family:Arial,sans-serif;font-size:13px;color:#222'>"
            f"<h2 style='color:#1F3864'>Back-Office Fee Remittance — {month_label(ym)}</h2>"
            f"<p>{len(bills)} bill(s), total <b>${total:,.2f}</b>. Spreadsheets updated, PDFs issued.</p>"
            f"<table style='border-collapse:collapse' cellpadding='6' border='1'>"
            f"<tr style='background:#1F3864;color:#fff'><th>Vendor</th><th>Bill No</th><th>Amount</th></tr>{lines}</table>"
            f"{status}</div>")
    subj = f"{tag} Back-Office Fee Remittance — {month_label(ym)} — {len(bills)} bills, ${total:,.2f}"
    # When already sent, don't re-attach PDFs (operator already has them via CC); manifest only.
    atts = [manifest_path] if already_sent else [manifest_path] + [b["Attachment (PDF)"] for b in bills]
    send_email(subject=subj, html=html, to=[NOTIFY_EMAIL], attachments=atts)
    print(f"   SUMMARY -> {NOTIFY_EMAIL} | {len(atts)} attachment(s)")

def email_bill_to_ramp(b):
    # One email = one bill (Ramp AP inbox OCRs one invoice per email). CC the
    # operator so they get a copy of every bill filed — no forwarding needed.
    subj = f"Bill: {b['Vendor (legal entity)']} — {b['Bill / Invoice No']} — ${b['Amount (USD)']:,.2f}"
    html = (f"<div style='font-family:Arial,sans-serif;font-size:13px'>"
            f"<p>Vendor: <b>{b['Vendor (legal entity)']}</b><br>Bill No: {b['Bill / Invoice No']}<br>"
            f"Amount: <b>${b['Amount (USD)']:,.2f}</b><br>Invoice date: {b['Invoice Date']} (due on receipt)</p>"
            f"<p>{b['Memo']}</p>"
            f"<p style='color:#888;font-size:11px'>Back-office booking fee remittance. Attached PDF is the remittance statement.</p></div>")
    send_email(subject=subj, html=html, to=[RAMP_BILLS_EMAIL], cc=[NOTIFY_EMAIL],
               attachments=[b["Attachment (PDF)"]])
    print(f"   SEND -> {RAMP_BILLS_EMAIL} (cc {NOTIFY_EMAIL}) | {subj}")

# ---------------- main ----------------
def prev_month():
    t = date.today(); y, m = (t.year, t.month - 1) if t.month > 1 else (t.year - 1, 12)
    return f"{y}-{m:02d}"

def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    send = "--send" in sys.argv
    notify = "--notify" in sys.argv
    ym = args[0] if args else prev_month()
    year, month = int(ym[:4]), int(ym[5:7])
    issue_date = date.today().strftime("%B %d, %Y")
    mode = "SEND (email bills to Ramp)" if send else ("NOTIFY (email summary to operator)" if notify else "DRY-RUN")
    print(f"== Back-Office Fee Remittance · target {month_label(ym)} · {mode} ==")
    tok = token()
    data, months = gather(year, month, tok)

    bills = []
    for ent in ENTITIES:
        wb_path, ytd = update_workbook(ent, data, months)
        rows = [(disp, data[slug][ym]["bookings"], data[slug][ym]["retail"], data[slug][ym]["fee"])
                for disp, slug in ent["locations"]]
        total = round(sum(r[3] for r in rows), 2)
        print(f"\n{ent['name']}: {month_label(ym)} = ${total:,.2f} | YTD ${ytd:,.2f} | wb {os.path.basename(wb_path)}")
        if total <= 0:
            print("   $0 — no PDF / no bill this month"); continue
        d = os.path.join(OUTBASE, ent["name"]); os.makedirs(d, exist_ok=True)
        pdf = os.path.join(d, f"{ent['name']} - {month_label(ym)} - Back-Office Fee Remittance.pdf")
        ok = render_pdf(build_html(ent, ym, rows, total, issue_date), pdf)
        print(f"   PDF {'OK' if ok else 'FAIL'}: {os.path.basename(pdf)}")
        if not ok:
            sys.exit(f"PDF render failed for {ent['name']} — is CHROMIUM_BIN correct ({CHROME})?")
        bills.append({"Vendor (legal entity)": ent["name"],
                      "Locations covered": "; ".join(dd for dd, _ in ent["locations"]),
                      "Bill / Invoice No": f"BOFR-{ent['code']}-{ym.replace('-','')}",
                      "Invoice Date": date.today().isoformat(), "Due Date": date.today().isoformat(),
                      "Terms": "Due on receipt", "Currency": "USD", "Amount (USD)": total,
                      "Memo": f"Peek back-office booking fee remittance — {month_label(ym)}",
                      "GL Account / Category": "", "Attachment (PDF)": pdf})

    if not bills:
        print("\nNo bills this month (all $0). Nothing to email.")
        return
    mpath = write_manifest(bills, ym)
    print(f"\nManifest: {os.path.basename(mpath)}  ({len(bills)} bills, ${sum(b['Amount (USD)'] for b in bills):,.2f})")
    if send:
        print("Emailing each bill to Ramp bills inbox (CC operator):")
        for b in bills: email_bill_to_ramp(b)
        notify_summary(bills, mpath, ym, already_sent=True)
    elif notify:
        print("Emailing review summary to operator (nothing sent to Ramp):")
        notify_summary(bills, mpath, ym, already_sent=False)
    else:
        print("[dry-run] --notify = email operator only · --send = email bills to Ramp (CC operator)")

if __name__ == "__main__":
    main()
